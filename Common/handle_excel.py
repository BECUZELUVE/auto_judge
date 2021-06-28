#!/usr/bin/env python
# -*- coding:utf-8 -*-
'''
@Project -> File   ：003auto_judge -> handle_excel.py
@IDE    ：PyCharm
@Author ：Icey
@Desc   ：
'''
from openpyxl import load_workbook


class HandleExcel:
    def __init__(self,file_path,sheet_name):
        self.wb = load_workbook(file_path)
        self.sh = self.wb[sheet_name]

    def read_titles(self):
        titles = []
        for item in list(self.sh.rows)[0]:
            titles.append(item.value)
        return titles

    def read_all_datas(self):
        all_datas = []
        titles = self.read_titles()

        for i in list(self.sh.rows)[1:]:
            values = []
            for c in i:
                values.append(c.value)
            res = dict(zip(titles, values))
            all_datas.append(res)

        return all_datas

    def close_file(self):
        self.wb.close()

    def output_excel(self,info,out_file_path):
        self.wb = load_workbook(out_file_path)
        # print(self.wb.sheetnames)
        if "P3考生成绩明细" in self.wb.sheetnames:
            del self.wb["P3考生成绩明细"]
        # print(self.wb.sheetnames)
        self.sh = self.wb.create_sheet("P3考生成绩明细")
        title = list(info[0].keys())
        self.sh.append(title)
        for i in range(len(info)):
            self.sh.append(list(info[i].values()))


        self.wb.save(out_file_path)







if __name__ == '__main__':
    import os
    from Common.handle_path import datas_dir
    file_path = os.path.join(datas_dir, "answer_datas.xlsx")
    out_file_path = os.path.join(datas_dir, "output_datas.xlsx")
    excel = HandleExcel(file_path,'P3考生成绩明细')
    cases = excel.read_all_datas()
    # for case in cases:
    #     print(case["第21题"])
    info = [{"id":1,"1":2,"2":3}]

    excel.output_excel(info,out_file_path)
    excel.close_file()